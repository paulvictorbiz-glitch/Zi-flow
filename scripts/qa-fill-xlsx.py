#!/usr/bin/env python3
"""
qa-fill-xlsx.py — writes Playwright auto-run verdicts into the QA tracker.

Reads scripts/qa-results.json (from qa-autorun.mjs), opens the blank
QA-Debug-Tracker-FootageBrain.xlsx, and for each result finds the row whose
Test ID (col A) matches on the named sheet, then writes:
  - Status (col H)
  - appends evidence to Notes (col J)
  - Tester (col L) = "Playwright auto"
  - Date tested (col M)
ONLY into cells that are currently empty / "Not started" — never overwrites
anything a human already filled. The Per-Role Access Matrix uses its own layout
(Status in col I, Notes in col J).

Saves to QA-Debug-Tracker-FootageBrain-FILLED.xlsx (template untouched).
Run:  python scripts/qa-fill-xlsx.py [YYYY-MM-DD]
"""

import json
import os
import sys

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "QA-Debug-Tracker-FootageBrain.xlsx")
OUT = os.path.join(ROOT, "QA-Debug-Tracker-FootageBrain-FILLED.xlsx")
RESULTS = os.path.join(ROOT, "scripts", "qa-results.json")
DATE = sys.argv[1] if len(sys.argv) > 1 else "2026-06-24"

# Per-sheet column layout for (status_col, notes_col, tester_col, date_col, id_col)
STD = dict(id=1, status=8, notes=10, tester=12, date=13)            # feature sheets
MATRIX = dict(id=1, status=9, notes=10, tester=None, date=None)     # Per-Role Access Matrix
SIMPLE_RT = dict(id=1, status=6, notes=7, tester=None, date=None)   # Realtime & Persistence
SIMPLE_IN = dict(id=1, status=7, notes=9, tester=None, date=None)   # Integration Health

SHEET_LAYOUT = {
    "Per-Role Access Matrix": MATRIX,
    "Realtime & Persistence": SIMPLE_RT,
    "Integration Health": SIMPLE_IN,
}

EMPTY = (None, "", "Not started")


def main():
    if not os.path.exists(RESULTS):
        print(f"!! {RESULTS} not found — run scripts/qa-autorun.mjs first.")
        sys.exit(1)
    with open(RESULTS, encoding="utf-8") as f:
        results = json.load(f)

    wb = load_workbook(SRC)
    # Build {sheet: {test_id: row}} index
    index = {}
    for ws in wb.worksheets:
        layout = SHEET_LAYOUT.get(ws.title, STD)
        idc = layout["id"]
        rowmap = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idc).value
            if v:
                rowmap[str(v).strip()] = r
        index[ws.title] = rowmap

    filled = skipped_missing = skipped_taken = 0
    per_sheet = {}
    unmatched = []

    for res in results:
        sheet = res["sheet"]
        tid = res["id"]
        status = res["status"]
        note = res.get("note", "")
        if sheet not in wb.sheetnames:
            unmatched.append((sheet, tid, "sheet-missing"))
            continue
        ws = wb[sheet]
        layout = SHEET_LAYOUT.get(sheet, STD)
        row = index.get(sheet, {}).get(tid)
        if not row:
            unmatched.append((sheet, tid, "id-missing"))
            skipped_missing += 1
            continue
        sc = ws.cell(row=row, column=layout["status"])
        if sc.value not in EMPTY:
            skipped_taken += 1   # human already filled — leave it
            continue
        sc.value = status
        nc = ws.cell(row=row, column=layout["notes"])
        nc.value = (str(nc.value) + " | " if nc.value else "") + note
        if layout["tester"]:
            ws.cell(row=row, column=layout["tester"]).value = "Playwright auto"
        if layout["date"]:
            ws.cell(row=row, column=layout["date"]).value = DATE
        filled += 1
        per_sheet[sheet] = per_sheet.get(sheet, 0) + 1

    wb.save(OUT)

    by = {}
    for res in results:
        by[res["status"]] = by.get(res["status"], 0) + 1
    print(f"Wrote {OUT}")
    print(f"Results: {len(results)}  ·  " + "  ".join(f"{k} {v}" for k, v in sorted(by.items())))
    print(f"Filled {filled} cells  ·  skipped {skipped_taken} (already filled)  ·  {skipped_missing} id-not-found")
    print("Per sheet filled:")
    for s, n in sorted(per_sheet.items()):
        print(f"  - {s}: {n}")
    if unmatched:
        print("Unmatched ids (check mapping):")
        for s, t, why in unmatched:
            print(f"  - {s} / {t} ({why})")


if __name__ == "__main__":
    main()
