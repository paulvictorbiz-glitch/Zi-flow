#!/usr/bin/env python3
"""
build-qa-tracker.py — generates QA-Debug-Tracker-FootageBrain.xlsx

A pre-populated, exhaustive QA / debugging tracker for the FootageBrain dashboard.
One sheet per tab/feature area, with dropdown data-validation, color-coded status,
a rollup Dashboard, a Per-Role Access Matrix, Realtime/Persistence + Integration
sheets, and a Bug Log.

Run:  python scripts/build-qa-tracker.py
Out:  QA-Debug-Tracker-FootageBrain.xlsx  (project root)

Pure local file generation — touches no app code, no DB, no network.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Styling constants
# --------------------------------------------------------------------------- #
HEADER_FILL   = PatternFill("solid", fgColor="1F2937")   # slate-800
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(bold=True, size=16, color="111827")
SUB_FONT      = Font(italic=True, size=10, color="6B7280")
SECTION_FONT  = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL  = PatternFill("solid", fgColor="374151")   # slate-700
WRAP_TOP      = Alignment(wrap_text=True, vertical="top")
TOP           = Alignment(vertical="top")
THIN          = Side(style="thin", color="D1D5DB")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Color rules for the Status column
PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")   # green
PASS_FONT    = Font(color="006100")
FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")   # red
FAIL_FONT    = Font(color="9C0006")
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")   # amber
WARN_FONT    = Font(color="9C6500")
GREY_FILL    = PatternFill("solid", fgColor="E5E7EB")   # grey
GREY_FONT    = Font(color="6B7280")

# --------------------------------------------------------------------------- #
# Schema
# --------------------------------------------------------------------------- #
HEADERS = [
    "Test ID", "Feature / Component", "What to check (test case)", "Steps",
    "Expected result", "Role(s)", "Priority", "Status", "Severity (if fail)",
    "Actual result / Notes", "Bug ID", "Tester", "Date tested",
]
WIDTHS = [11, 28, 46, 40, 40, 16, 9, 13, 14, 44, 10, 12, 13]
WRAP_COLS = {3, 4, 5, 10}            # 1-based: What/Steps/Expected/Notes
STATUS_OPTS   = '"Not started,Pass,Fail,Blocked,N-A,Retest"'
SEVERITY_OPTS = '"Critical,Major,Minor,Cosmetic"'
PRIORITY_OPTS = '"P0,P1,P2,P3"'

STATUS_COL   = 8    # H
SEVERITY_COL = 9    # I
PRIORITY_COL = 7    # G


def style_header(ws, row=1, ncols=len(HEADERS)):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def add_feature_sheet(wb, title, rows):
    """rows = list of dicts with keys: id, feat, check, steps, exp, role, prio.
    Optionally a row can be {'section': 'Heading'} to insert a band."""
    ws = wb.create_sheet(title=title[:31])
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
    style_header(ws)

    r = 2
    for row in rows:
        if "section" in row:
            ws.cell(row=r, column=1, value=row["section"])
            ws.cell(row=r, column=1).font = SECTION_FONT
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=c).fill = SECTION_FILL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
            r += 1
            continue
        vals = [
            row.get("id", ""), row.get("feat", ""), row.get("check", ""),
            row.get("steps", ""), row.get("exp", ""), row.get("role", "all"),
            row.get("prio", "P2"), "Not started", "", "", "", "", "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_TOP if c in WRAP_COLS else TOP
            cell.border = BORDER
        r += 1

    last = r - 1
    # Freeze header, autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(last,1)}"

    if last >= 2:
        # Dropdowns
        dv_status = DataValidation(type="list", formula1=STATUS_OPTS, allow_blank=True)
        dv_sev    = DataValidation(type="list", formula1=SEVERITY_OPTS, allow_blank=True)
        dv_prio   = DataValidation(type="list", formula1=PRIORITY_OPTS, allow_blank=True)
        ws.add_data_validation(dv_status); ws.add_data_validation(dv_sev); ws.add_data_validation(dv_prio)
        sl, sc = get_column_letter(STATUS_COL), get_column_letter(SEVERITY_COL)
        pl = get_column_letter(PRIORITY_COL)
        dv_status.add(f"{sl}2:{sl}{last}")
        dv_sev.add(f"{sc}2:{sc}{last}")
        dv_prio.add(f"{pl}2:{pl}{last}")
        # Conditional formatting on Status
        rng = f"{sl}2:{sl}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL, font=PASS_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL, font=FAIL_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=WARN_FILL, font=WARN_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Retest"'], fill=WARN_FILL, font=WARN_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N-A"'], fill=GREY_FILL, font=GREY_FONT))
    return ws, last


# --------------------------------------------------------------------------- #
# Row helpers
# --------------------------------------------------------------------------- #
def R(id, feat, check, steps, exp, role="all", prio="P2"):
    return dict(id=id, feat=feat, check=check, steps=steps, exp=exp, role=role, prio=prio)

def S(label):
    return dict(section=label)


# =========================================================================== #
#  TEST DATA  — one list per sheet
# =========================================================================== #

AUTH = [
    S("Sign-in & session"),
    R("AUTH-001", "Sign-in screen", "Valid email + password signs in", "Open app logged out → enter valid creds → submit", "Lands on dashboard (/app), correct identity loaded", "all", "P0"),
    R("AUTH-002", "Sign-in screen", "Wrong password is rejected with a clear error", "Enter valid email + wrong password → submit", "Inline error shown; no crash; stays on sign-in", "all", "P1"),
    R("AUTH-003", "Sign-in screen", "No self-signup path exists", "Inspect sign-in screen for any 'Create account' link", "No registration option (owner-only via admin)", "all", "P2"),
    R("AUTH-004", "Claim-Identity screen", "Auth user with no people slot sees claim screen", "Sign in as an auth user not bound to a people row", "ClaimIdentityScreen shows 'ask Paul' message; no dashboard", "all", "P1"),
    R("AUTH-005", "Splash / gates", "Loading splashes appear then resolve", "Sign in on a cold load", "'signing in…' / 'loading identity…' show briefly, then app", "all", "P3"),
    R("AUTH-006", "Session persistence", "Session survives reload", "Sign in → hard refresh", "Still signed in, no re-prompt", "all", "P1"),
    R("AUTH-007", "Sign-out", "Sign-out returns to sign-in", "Use sign-out control", "Back to sign-in screen; protected routes blocked", "all", "P1"),
    R("AUTH-008", "Landing (/)", "Public landing renders without auth", "Open / while logged out", "Landing page loads, no console errors", "all", "P2"),
]

NAV = [
    S("Tab visibility (run per role — see Per-Role Access Matrix for the grid)"),
    R("NAV-001", "Nav drawer", "Menu opens and lists groups/tabs", "Click Menu", "Drawer opens; 7 groups; only allowed tabs shown", "all", "P0"),
    R("NAV-002", "All tabs render", "Each visible tab opens without blank/error", "Click every tab once (smoke)", "Each renders; no console error; no white screen", "all", "P0"),
    R("NAV-003", "Lazy load fallback", "Heavy tabs show loading fallback then render", "Open a lazy tab (Editor, Monitor, Training…)", "'loading…' shows briefly, chunk loads, renders", "all", "P2"),
    R("NAV-004", "Safety-net bounce", "Navigating to a hidden tab bounces to allowed", "As non-owner, force view=monitor (URL/history)", "Auto-redirects to first allowed tab, not 404", "skilled", "P1"),
    R("NAV-005", "Nav group reorder", "Drag group header reorders + persists", "Drag a group header; reload", "Order changes; persists (nav_group_order)", "all", "P3"),
    R("NAV-006", "Group expand/collapse", "Groups collapse/expand; Escape closes drawer", "Toggle a group; press Escape", "Collapse state holds; Escape closes drawer", "all", "P3"),
    S("Deep-links & aliases"),
    R("NAV-007", "Deep-link ?reel=", "?reel=REEL-301 opens that detail", "Visit /app?reel=<id>", "Opens detail for the reel; query stripped after", "all", "P1"),
    R("NAV-008", "Deep-link ?compare=1", "?reel=&compare=1 opens compare", "Visit /app?reel=<id>&compare=1", "Detail opens in compare mode", "all", "P2"),
    R("NAV-009", "Share-target ?capture=", "?capture=1&url= prefills Reel DNA", "Visit /app?capture=1&url=<link>", "Reel DNA form prefilled; query stripped", "all", "P2"),
    R("NAV-010", "Legacy wb_view alias", "Old pulse/ai view redirects to monitor", "Set localStorage wb_view=pulse; reload", "Lands on Monitor, not a dead view", "owner", "P3"),
    R("NAV-011", "Breadcrumb + live status", "Breadcrumb tracks view; 'realtime · live' shows", "Navigate between tabs and a reel detail", "Breadcrumb updates; live indicator present", "all", "P3"),
    R("NAV-012", "Needs-you badge", "Red dot on Menu when work awaits", "Have a reel awaiting you / unread inbox", "Badge shows correct count; clears when resolved", "all", "P2"),
]

MYWORK = [
    R("MYW-001", "Assigned reels list", "Shows reels assigned to the logged-in person", "Open My work as each role", "Only this person's relevant reels appear", "all", "P1"),
    R("MYW-002", "'Needs you' count", "Badge count matches reels awaiting action", "Compare badge to actual pending reels", "Count is accurate per ownership + stage", "all", "P1"),
    R("MYW-003", "Review-queue surfacing", "Reviewer sees reels in review", "Login as reviewer with reels in Review", "Review items surface here", "reviewer", "P1"),
    R("MYW-004", "Recent messages card", "Recent Rocket.Chat messages appear", "Have recent team-chat messages", "Card lists recent messages (cap ~30)", "all", "P2"),
    R("MYW-005", "Open reel from card", "Clicking a reel opens detail", "Click a reel card", "Detail view opens for that reel", "all", "P1"),
]

PIPELINE = [
    S("Filters & layout"),
    R("PIPE-001", "Filter pills", "'All reels' / 'Blocked' filter toggles", "Click Blocked/warn pill", "List narrows to blocked; pill highlights; toggle off restores", "all", "P2"),
    R("PIPE-002", "Group by series", "Series clustering toggles", "Toggle Group by Series", "Reels cluster by series within cells", "all", "P2"),
    R("PIPE-003", "Card view switcher", "List / 2×2 / 3×3 change card density", "Click each of the 3 view buttons", "Card layout changes accordingly", "all", "P3"),
    R("PIPE-004", "Columns menu", "Hide/show stage columns + persist", "Uncheck a stage; 'Show all columns'; reload", "Columns hide/show; hidden count shown; persists", "all", "P2"),
    R("PIPE-005", "Lanes menu", "Hide/show team lanes + persist (Supabase)", "Uncheck a lane; reload (and in 2nd browser)", "Lane hides; count amber; persists per-user", "all", "P2"),
    S("Drag, drop & multi-select"),
    R("PIPE-006", "Drag card between stages", "Move a card to a new stage/lane", "Drag a card to another cell", "Card moves; stage/lane updates; persists", "all", "P0"),
    R("PIPE-007", "Card reorder", "Drop before/after another card", "Drop onto another card", "Order changes; board_order persists", "all", "P2"),
    R("PIPE-008", "Multi-select", "Cmd/Ctrl/Shift-click selects multiple", "Modifier-click several cards", "Chip shows count + 'drag any to move group'", "all", "P2"),
    R("PIPE-009", "Group drag", "Dragging one moves all selected", "Select 3, drag one to a new cell", "All 3 move together", "all", "P2"),
    R("PIPE-010", "Clear selection", "'Clear' link deselects; plain click opens detail", "Click Clear; then plain-click a card", "Selection clears; plain click opens detail", "all", "P3"),
    R("PIPE-011", "Drop to Posted modal", "Schedule date modal on drop to Posted", "Drag a card to Posted", "Date picker modal; Move/Cancel work; date saved", "all", "P1"),
    R("PIPE-012", "Completed-drop block", "Non-owner blocked from Completed; header flashes", "As editor, drag to Completed", "Header flashes red 700ms; drop rejected", "skilled", "P2"),
    R("PIPE-013", "Lane right-click hide", "Owner hides a lane via context menu", "Right-click a lane → Hide this lane", "Lane hides; DB updated", "owner", "P3"),
    S("Sub-modes"),
    R("PIPE-014", "Board/List/Calendar/Archived", "Sub-mode bar switches views + persists", "Switch each sub-mode; reload", "Each view renders; wb_pipeline_mode persists", "all", "P1"),
    R("PIPE-015", "List view lanes", "List shows lanes + counts", "Open List view", "Lanes with counts; toggles honored", "all", "P2"),
    R("PIPE-016", "Calendar view", "Calendar renders scheduled reels", "Open Calendar view", "Scheduled reels appear on correct dates", "all", "P2"),
    R("PIPE-017", "Archived view", "Archived reels listed + restorable", "Open Archived view", "Archived reels show; restore works", "all", "P2"),
]

REELDNA = [
    S("Capture"),
    R("DNA-001", "Capture form", "URL paste detects platform", "Paste IG/TikTok/YT link", "Platform auto-detected; pill set", "all", "P1"),
    R("DNA-002", "Capture form", "Genes + notes + Capture saves a row", "Pick genes, add notes, click Capture", "New DNA row appears with fields", "all", "P0"),
    R("DNA-003", "Capture form", "Capture disabled when URL empty; Enter submits", "Empty URL; then valid + Enter key", "Button disabled empty; Enter submits", "all", "P3"),
    R("DNA-004", "Bookmarklet", "Drag-to-bookmark prefills capture", "Use bookmarklet on an external page", "Returns to app with form prefilled", "all", "P3"),
    S("Spreadsheet — filters & marks"),
    R("DNA-005", "Column filters", "Per-column text filter narrows rows live", "Type in a column 'filter…' box", "Rows filter as you type; no focus loss", "all", "P1"),
    R("DNA-006", "Clear column filters", "✕ clears all column filters", "Set filters → click ✕", "All filters cleared", "all", "P3"),
    R("DNA-007", "Status filter", "Status dropdown filters rows", "Pick a status", "Only that status shows; 'All' resets", "all", "P2"),
    R("DNA-008", "Row star", "☆/★ toggles favorite; header star filters", "Star a row; toggle header star filter", "Favorite persists; filter shows only starred", "all", "P2"),
    R("DNA-009", "Row color dot", "Color popover tints row; × clears", "Open dot popover; pick a color; clear", "Row tinted; persists; header color filter works", "all", "P2"),
    S("Spreadsheet — editing & actions"),
    R("DNA-010", "Inline cell edit", "Edit Location/Music/Font/SFX/Story/Notes", "Click a cell, type, Enter/blur; Escape cancels", "Saves on commit; Escape cancels; tag-notes promote", "all", "P1"),
    R("DNA-011", "Status dropdown (row)", "Per-row status select updates DB", "Change a row status", "reel_dna.status updates", "all", "P2"),
    R("DNA-012", "Assets button", "▣ opens Reel Assets page", "Click the assets button", "Full-screen assets page opens", "all", "P2"),
    R("DNA-013", "Row actions", "Card/Compare/DNA/Deconstructor open", "Click ⤢ / ⇔ / DNA / ▦ per row", "Each opens the right view", "all", "P2"),
    R("DNA-014", "→ Pipeline", "Send-to-pipeline editor picker fans out", "Click → Pipeline; select editors; send", "One copy per editor in Not Started; link shows", "all", "P1"),
    R("DNA-015", "↩ DNA", "Pull a reel back from pipeline", "Click ↩ DNA on a linked row", "Reel returns to DNA; link cleared", "all", "P2"),
    R("DNA-016", "Archive / Delete", "⧉ archive toggle and ✕ delete", "Archive then restore; delete a test row", "Archive toggles; delete removes permanently", "all", "P2"),
    R("DNA-017", "Show Sent / Archived", "Toggles include sent/archived rows", "Toggle Show Sent; Show Archived", "Rows include/exclude; archive count shown", "all", "P3"),
    S("IG Sync"),
    R("DNA-018", "Refresh poller", "Refresh triggers IG poll + reloads", "Click Refresh", "'Refreshing…' state; rows reload after delay", "owner", "P1"),
    R("DNA-019", "Check IG Sync", "Health report loads run history + issues", "Click Check IG Sync → Show report", "IgSyncHealth panel shows landed/coverage/errors", "owner", "P2"),
]

THUMB = [
    R("THM-001", "Capture form", "YouTube URL → video-id + thumbnail preview", "Paste a YT link", "Video id extracted; thumbnail preview (maxres→hq)", "all", "P1"),
    R("THM-002", "Capture", "Genes + notes + Capture saves card", "Pick genes, capture", "New thumbnail card appears", "all", "P1"),
    R("THM-003", "View switcher", "Cards ↔ Spreadsheet toggle", "Toggle view mode", "Both views render same data", "all", "P2"),
    R("THM-004", "Gene editing", "Inline gene edit (both views)", "Edit a gene; blur/Escape", "Saves on commit; Escape cancels", "all", "P2"),
    R("THM-005", "Status chips", "Change status via chips", "Click a status chip", "Status updates", "all", "P2"),
    R("THM-006", "Status filter + count", "Filter by status with counts", "Use status filter pills", "Rows filter; counts correct", "all", "P3"),
    R("THM-007", "Archive / Delete", "Archive toggle + delete per card/row", "Archive then delete a test card", "Archive toggles; delete removes", "all", "P2"),
    R("THM-008", "Refresh", "Playlist poller reload (if configured)", "Click Refresh", "thumbnail_dna reloads", "owner", "P3"),
]

LIBS = [
    S("Footage"),
    R("LIB-001", "Footage library", "Search returns clips", "Open Footage; search a term", "Matching clips listed", "all", "P2"),
    R("LIB-002", "Footage attach/detach", "Attach a clip then remove it", "Attach to a reel; detach", "Attach + detach persist", "all", "P2"),
    R("LIB-003", "Footage thumbnails", "Drive/footage thumbnails render", "Browse footage with Drive sources", "Thumbnails load (no broken images)", "all", "P2"),
    S("Coverage"),
    R("LIB-004", "Coverage page", "Coverage tracking renders", "Open Coverage", "Page loads; data present", "all", "P3"),
    S("Locations"),
    R("LIB-005", "Locations map", "Map + locations render (LocationsProvider)", "Open Locations", "Map loads; markers present; no crash", "all", "P3"),
]

DETAIL = [
    R("DET-001", "Footage Brain search", "Search + attach footage to reel", "In detail, search then attach", "Clip attaches to reel", "all", "P1"),
    R("DET-002", "Attached footage list", "Detach removes a clip", "Remove an attached clip", "Clip detaches (if removeFootage allowed)", "all", "P2"),
    R("DET-003", "Logline editor", "Edit logline (owner) / read-only (editors)", "Edit logline as owner, then as editor", "Owner saves; editor sees read-only by default", "all", "P1"),
    R("DET-004", "Script editor", "Beat plan textarea saves on blur", "Edit script; blur; Escape cancels", "Saves; Escape cancels", "all", "P2"),
    R("DET-005", "Voiceover editor", "Voiceover field saves / gated", "Edit voiceover per role", "Owner saves; editor gated by editVoiceover", "all", "P2"),
    R("DET-006", "Comments", "Add comment + thread renders", "Post a comment", "Comment appears with author/timestamp", "all", "P1"),
    R("DET-007", "Music Picker modal", "Attach a track to the reel", "Open music picker; add track", "Track attaches to reel assets", "all", "P2"),
    R("DET-008", "Share to channel", "Share reel reference to Rocket.Chat", "Use share-to-channel", "Reference card posts to chat", "all", "P2"),
    R("DET-009", "Color picker", "5-swatch card color (gated)", "Change card color per role", "Owner changes; editor gated by changeCardColor", "all", "P3"),
]

EDITOR = [
    S("Embed & UI preset"),
    R("EDT-001", "Editor embed loads", "OpenCut iframe loads in Editor tab", "Open Editor tab", "iframe (editor.footagebrain.com) loads, no CSP/frame error", "all", "P0"),
    R("EDT-002", "CapCut default", "Editor defaults to CapCut UI", "Open Editor fresh", "CapCut skin shown by default", "all", "P1"),
    R("EDT-003", "CapCut↔Classic toggle", "Toggle switches + persists per-user", "Toggle to Classic; reload; reopen", "View switches; preset persists (editor_ui_preset)", "all", "P1"),
    R("EDT-004", "Classic explicit", "?ui=classic forces classic", "Open editor with ?ui=classic", "Classic OpenCut shown", "all", "P3"),
    S("Editing engine"),
    R("EDT-005", "Playback / scrub", "Play, pause, scrub timeline", "Load a project; play; scrub", "Playhead advances; pause works; scrub smooth", "all", "P0"),
    R("EDT-006", "Add clips", "Add clips from attached footage", "Add a clip to timeline", "Clip appears on a track", "all", "P1"),
    R("EDT-007", "+ Add music", "Music picker integrates", "Click +Add music", "Music library opens; track adds to audio track", "all", "P2"),
    R("EDT-008", "Render draft", "Editor renders a draft", "Submit a render; poll status", "Render job runs; draft output produced", "all", "P1"),
    R("EDT-009", "Auto-captions", "Captions fill text track (if enabled)", "Run auto-captions", "Text track populates progressively", "all", "P2"),
    R("EDT-010", "Silence/filler trim", "Detect + accept/reject silence ranges", "Run trim; accept a range", "Ranges highlighted; apply splits clip", "all", "P3"),
    R("EDT-011", "Command palette", "Ctrl+Shift+P opens palette", "Press Ctrl+Shift+P", "Command palette opens; search works", "all", "P3"),
    R("EDT-012", "Shortcuts editor", "View/rebind keybindings persists", "Open shortcuts dialog; rebind a key", "Rebind saves; conflict detection works", "all", "P3"),
]

EDITOR_COLLAB = [
    S("Projects browser"),
    R("ECP-001", "Projects list", "Projects tab lists projects", "Open Projects", "Project cards render with status badges", "all", "P1"),
    R("ECP-002", "New blank project", "Create empty project", "New blank → title → create", "Empty project opens in editor", "all", "P1"),
    R("ECP-003", "New from reel", "Create project from a reel", "New from reel → pick reel", "Project titled from reel; opens", "all", "P2"),
    R("ECP-004", "⋯ Rename", "Rename a project", "⋯ → Rename → new name", "Name updates on card", "all", "P2"),
    R("ECP-005", "⋯ Archive", "Archive/Unarchive a project", "⋯ → Archive; then Unarchive", "Badge toggles archived/draft", "all", "P2"),
    R("ECP-006", "⋯ Delete", "Delete a project (confirm)", "⋯ → Delete → confirm", "Project removed after confirm", "all", "P2"),
    R("ECP-007", "Menu portal", "⋯ menu not clipped by overflow", "Open ⋯ near edge of scroll area", "Menu fully visible (portaled)", "all", "P3"),
    S("Collaboration (two sessions)"),
    R("ECP-008", "Take control", "First editor takes the single-writer lock", "Open same project in 2 browsers; Take control", "Holder can edit; lock acquired (30s, heartbeat 10s)", "all", "P1"),
    R("ECP-009", "Lock badge", "Viewers see '🔒 name editing' + read-only", "Observe 2nd browser while holder edits", "2nd is read-only; shows holder name", "all", "P1"),
    R("ECP-010", "Timeline sync", "Holder edits stream to viewers", "Holder changes timeline; watch viewer", "Viewer timeline updates in realtime", "all", "P1"),
    R("ECP-011", "Release control", "Release frees the lock", "Holder clicks Release", "Lock freed; other can take control", "all", "P2"),
    R("ECP-012", "Lock expiry", "Lock auto-frees when holder idle/closes", "Holder closes tab; wait", "Lock expires; control available", "all", "P2"),
    R("ECP-013", "Presence avatars", "Per-person avatar + playhead shown", "Both scrub in same project", "Colored pills; playheads update; ghost-hide 60s", "all", "P2"),
]

EXPORT = [
    S("Accounts & selection (owner)"),
    R("EXP-001", "Accounts dropdown", "Per-platform + select-all accounts", "Open accounts dropdown", "Platforms listed; select-all works; closes on outside click", "owner", "P1"),
    R("EXP-002", "Reel selection", "Per-row + select-all (indeterminate) checkboxes", "Select reels; use header select-all", "Selection tracks; indeterminate state correct", "owner", "P2"),
    S("Push to Planable"),
    R("EXP-003", "Push button state", "Disabled until reels + accounts chosen", "Try push with none selected", "Button disabled; label shows N×M when ready", "owner", "P2"),
    R("EXP-004", "Preview modal", "Preview lists per-reel target/caption/schedule", "Click Push → preview", "Cards show handle, caption, schedule, media flag", "owner", "P1"),
    R("EXP-005", "Posting time input", "HH:mm time applies to schedule", "Set posting time in preview", "Composed schedule uses chosen time per date", "owner", "P1"),
    R("EXP-006", "SKIP logic", "Past/blank-date reels flagged SKIP", "Include a past-dated reel", "Marked SKIP; not pushed", "owner", "P2"),
    R("EXP-007", "Confirm push", "Creates DRAFT posts (never auto-publish)", "Confirm push", "Drafts created in Planable; campaign indicator shows", "owner", "P0"),
    R("EXP-008", "Grouping", "Each reel = own grouped cross-page post", "Push 2 reels to multiple pages", "Each reel one groupId; batch under one campaign", "owner", "P1"),
    R("EXP-009", "Cancel", "Cancel closes without sending", "Open preview → Cancel", "No posts created", "owner", "P3"),
    S("CSV"),
    R("EXP-010", "Download CSV", "Export posted reels to CSV", "Click Download CSV", "RFC-4180 CSV downloads; disabled when none", "owner", "P3"),
]

MUSIC = [
    R("MUS-001", "Search view", "Debounced search returns tracks", "Type a query", "Results after ~350ms; relevant tracks", "all", "P2"),
    R("MUS-002", "Browse chips", "Genre/Mood chips search", "Click a genre and a mood chip", "Tracks filtered by chip", "all", "P3"),
    R("MUS-003", "Play preview", "Single audio plays/pauses across views", "Play one track, then another", "Only one plays; switching stops prior", "all", "P2"),
    R("MUS-004", "Favorite", "Toggle favorite per track", "Heart a track; open Favorites", "Track appears in Favorites; persists", "all", "P2"),
    R("MUS-005", "Download", "Licensed download resolves a URL", "Click Download", "'Downloading…' then a valid signed URL", "all", "P1"),
    R("MUS-006", "Playlists CRUD", "Create/rename/delete playlist", "Create a playlist; rename; delete", "All three persist", "all", "P3"),
    R("MUS-007", "Add/remove in playlist", "Add a track then remove it", "Add to playlist; remove", "Track count updates", "all", "P3"),
    R("MUS-008", "Attach to reel", "Attach a track to a reel/DNA asset", "Use Attach to reel", "Saved to reel assets", "all", "P2"),
]

SCOUT = [
    R("SCT-001", "Search", "Live filter by name/desc/dossier", "Type a query", "Rows filter live", "owner", "P2"),
    R("SCT-002", "Filters", "Source / category / score / fav / archived", "Apply each filter", "Rows narrow correctly; combine cleanly", "owner", "P2"),
    R("SCT-003", "Group by category", "Toggle groups rows with headers", "Toggle Group by Category", "Collapsible category groups appear", "owner", "P3"),
    R("SCT-004", "Column sorts", "Sort name/score/created/traction", "Click each sortable header", "Arrow indicator; order flips on re-click", "owner", "P2"),
    R("SCT-005", "Row star", "Star/unstar updates shortlist", "Star a row", "Favorite persists (Scout DB)", "owner", "P2"),
    R("SCT-006", "Row archive", "Archive/unarchive a product", "Toggle archive", "Archived state persists", "owner", "P3"),
    R("SCT-007", "Row delete", "Delete with confirm", "Delete a test row → confirm", "Row removed after confirm", "owner", "P2"),
    R("SCT-008", "Row expand", "Dossier detail expands", "Click a row", "Summary/target/tech/difficulty/angle/model show", "owner", "P2"),
    R("SCT-009", "Refresh scrape", "Refresh fires backend scraper", "Click Refresh", "Toast 'scraping started'; reload ~2min later", "owner", "P1"),
    R("SCT-010", "Reload", "Reload re-fetches products", "Click Reload", "Latest rows + dossiers load", "owner", "P2"),
]

MONITOR = [
    R("MON-001", "Sub-tab strip", "Infra/Pulse/AI/Scout sub-tabs gate by perm", "Open Monitor; switch sub-tabs", "Only permitted sub-tabs show; each renders", "owner", "P1"),
    R("MON-002", "Provider cards", "Supabase/Hetzner/GCloud usage + sparklines", "View Infra sub-tab", "Usage bars, sparklines; 80%+ amber, 95%+ red", "owner", "P1"),
    R("MON-003", "Threshold toast", "Toast fires when a metric crosses 80%", "Observe a high-usage metric", "Toast notification fires", "owner", "P2"),
    R("MON-004", "Last-good cache", "Stale provider keeps last-good value", "Provider API times out", "Last-good value retained, not blank", "owner", "P3"),
    R("MON-005", "Frontend perf card", "p75 load + INP indicator + sparkline", "View perf card", "p75 ms, INP color-coded, 7-day median line", "owner", "P2"),
    R("MON-006", "Editor usage monitor", "Live oc_locks + per-person history", "View editor-usage card", "Active locks + per-person/sparkline (mig 0097)", "owner", "P2"),
    R("MON-007", "Pulse sub-tab", "Pulse alerts render", "Open Pulse sub-tab", "Algorithm/news items render", "owner", "P3"),
    R("MON-008", "AI brain sub-tab", "FAQ/insights + Ask LLM", "Open AI sub-tab; ask a question", "FAQ list; insights; Ask LLM returns", "owner", "P3"),
]

TEAMCHAT = [
    R("TCH-001", "RC iframe loads", "chat.footagebrain.com embeds + stays mounted", "Open Team; switch away and back", "iframe persists (socket alive); no reload", "all", "P1"),
    R("TCH-002", "Reel-share search", "Search reel by id/title in picker", "Type in share picker", "Autocomplete matches (≤100); arrows/Enter select", "all", "P2"),
    R("TCH-003", "Send share", "Channel + feedback + Send posts card", "Pick reel + channel + feedback → Send", "Card posts to channel; comment saved on reel; toast", "all", "P1"),
    R("TCH-004", "New-message ping", "Audible ping on new message (unmuted)", "Coworker posts in chat", "Ping + toast; Team badge increments", "all", "P2"),
    R("TCH-005", "Mute toggle", "Mute silences pings + persists", "Toggle mute; reload", "No pings when muted; persists per-user", "all", "P2"),
    R("TCH-006", "Unseen badge", "Team tab badge counts unseen", "Receive messages while on another tab", "Badge shows correct unseen count", "all", "P2"),
    R("TCH-007", "Desktop notification", "Notification when window unfocused", "Blur window; receive a message", "OS notification; click focuses Team tab", "all", "P3"),
    R("TCH-008", "Cross-tab de-dupe", "Same message not double-pinged", "Two tabs open; one message", "Only one ping (seen-id ring)", "all", "P3"),
]

TRAINING = [
    R("TRN-001", "Module expand", "Modules expand/collapse", "Click a module title", "Expands; focusModule auto-scrolls", "all", "P2"),
    R("TRN-002", "Progress checkboxes", "Tick lessons; auto-complete module", "Check all lessons in a module", "Module marks complete", "all", "P2"),
    R("TRN-003", "Inline content edit", "Owner edits prose/lists; editors read-only", "Edit content as owner, view as editor", "Owner edits persist; editor read-only", "all", "P2"),
    R("TRN-004", "Quiz", "Multiple-choice + best score", "Take a quiz", "Scores; best score persists", "all", "P3"),
    R("TRN-005", "Flashcards", "Flip + progress through deck", "Flip cards", "Reveals answers; progresses", "all", "P3"),
    R("TRN-006", "Chapters", "Chapter nav scrolls within module", "Use chapter nav", "Scrolls to chapter; panels stay mounted", "all", "P3"),
    R("TRN-007", "Resources", "Resources link sheet renders + links work", "Open Resources", "Links present and open", "all", "P3"),
]

MISC = [
    S("Inbox"),
    R("MSC-001", "Inbox unread badge", "Badge counts unread comments/DMs", "Receive a comment/DM", "Badge increments; clears on read", "all", "P2"),
    R("MSC-002", "Inbox threads", "Comment/DM threads render across platforms", "Open Inbox", "Threads render; reply works (if supported)", "all", "P2"),
    S("Analytics"),
    R("MSC-003", "Analytics panels", "Analytics renders (live or mock)", "Open Analytics", "Panels load; no crash when unconnected", "all", "P3"),
    S("Generate"),
    R("MSC-004", "Idea Generator", "AI ideation generates results", "Open Generate; run", "Ideas returned (paid; gated)", "owner", "P3"),
    S("Lossless"),
    R("MSC-005", "Lossless cut", "In-browser lossless utility loads + cuts", "Open Lossless; load a clip; cut", "Utility loads; cut works", "all", "P3"),
]

ADMIN = [
    R("ADM-001", "Open Roles admin", "Owner-only admin page opens", "Avatar → Roles & permissions", "Admin matrix loads (owner only)", "owner", "P0"),
    R("ADM-002", "Toggle capability", "Toggle a view/action for a role", "Toggle a cap; Save", "Saves to app_settings + localStorage; effect applies", "owner", "P1"),
    R("ADM-003", "Per-person override", "Override a cap for one person", "Set a per-person toggle", "Overrides role default for that person", "owner", "P1"),
    R("ADM-004", "Perspective preview", "Owner previews a role's gated view", "Switch perspective in topbar", "UI gates to that role; reset on reload", "owner", "P1"),
    R("ADM-005", "Add / invite editor", "activate-slot creates auth user", "Add new editor → email → invite", "Account created (email_confirm); slot linked", "owner", "P1"),
    R("ADM-006", "Set password", "Owner resets a user's password", "Use set-password", "Password updated; user can sign in", "owner", "P2"),
    R("ADM-007", "Update email", "Owner updates a user's email", "Use update-email", "Email updated", "owner", "P2"),
    R("ADM-008", "Delete user", "Remove a person/auth user safely", "Delete a test user", "Slot removed; no cascade data loss", "owner", "P2"),
]

# Per-Role Access Matrix — its own column shape (not the standard feature schema)
ROLE_MATRIX_HEADERS = [
    "Check ID", "Tab / Action", "Type", "owner", "skilled", "variant",
    "reviewer", "demo", "Status", "Notes",
]
# (expected: V=Visible, H=Hidden, A=Allowed, X=Blocked)
ROLE_ROWS = [
    ("RM-01", "mywork", "tab", "V", "V", "V", "V", "V"),
    ("RM-02", "pipeline", "tab", "V", "V", "V", "V", "V"),
    ("RM-03", "reeldna", "tab", "V", "V", "V", "V", "V"),
    ("RM-04", "footage", "tab", "V", "V", "V", "V", "V"),
    ("RM-05", "training", "tab", "V", "V", "V", "V", "H"),
    ("RM-06", "resources", "tab", "V", "V", "V", "V", "H"),
    ("RM-07", "team", "tab", "V", "V", "V", "V", "H"),
    ("RM-08", "editor", "tab", "V", "H", "H", "H", "V"),
    ("RM-09", "projects", "tab", "V", "H", "H", "H", "V"),
    ("RM-10", "lossless", "tab", "V", "H", "H", "H", "V"),
    ("RM-11", "export", "tab", "V", "H", "H", "H", "V"),
    ("RM-12", "analytics", "tab", "V", "H", "H", "H", "V"),
    ("RM-13", "inbox", "tab", "V", "H", "H", "H", "V"),
    ("RM-14", "generate", "tab", "V", "H", "H", "H", "H"),
    ("RM-15", "music", "tab", "V", "H", "H", "H", "H"),
    ("RM-16", "coverage", "tab", "V", "H", "H", "H", "H"),
    ("RM-17", "locations", "tab", "V", "H", "H", "H", "H"),
    ("RM-18", "monitor", "tab", "V", "H", "H", "H", "H"),
    ("RM-19", "activity", "tab", "V", "H", "H", "H", "H"),
    ("RM-20", "settings (Roles admin)", "tab", "V", "H", "H", "H", "H"),
    ("RM-21", "createReel", "action", "A", "A", "A", "A", "A"),
    ("RM-22", "deleteReel", "action", "A", "X", "X", "X", "X"),
    ("RM-23", "archiveReel", "action", "A", "A", "A", "A", "A"),
    ("RM-24", "approveReview", "action", "A", "X", "X", "A", "X"),
    ("RM-25", "attachFootage", "action", "A", "A", "A", "A", "A"),
    ("RM-26", "moveReel", "action", "A", "A", "A", "A", "A"),
    ("RM-27", "moveToCompleted", "action", "A", "X", "X", "X", "X"),
    ("RM-28", "changeCardColor", "action", "A", "X", "X", "X", "X"),
    ("RM-29", "editLogline", "action", "A", "X", "X", "X", "X"),
    ("RM-30", "editScript", "action", "A", "X", "X", "X", "X"),
    ("RM-31", "editVoiceover", "action", "A", "X", "X", "X", "X"),
    ("RM-32", "removeFootage", "action", "A", "X", "X", "X", "X"),
    ("RM-33", "tagReelSkills", "action", "A", "X", "X", "X", "X"),
    ("RM-34", "editReelId", "action", "A", "X", "X", "X", "X"),
    ("RM-35", "bulkMoveReels", "action", "A", "X", "X", "X", "X"),
    ("RM-36", "gradeRubric", "action", "A", "X", "X", "A", "X"),
    ("RM-37", "editManual (training)", "action", "A", "X", "X", "X", "X"),
]

REALTIME_HEADERS = [
    "Test ID", "Category", "What to check", "How to test", "Expected",
    "Status", "Notes",
]
REALTIME_ROWS = [
    ("RT-01", "Realtime", "Move reel syncs across sessions", "Move a reel in browser A", "Browser B updates <500ms"),
    ("RT-02", "Realtime", "New reel appears in other session", "Create reel in A", "Appears in B"),
    ("RT-03", "Realtime", "New comment streams", "Comment on a reel in A", "B's detail updates if viewing"),
    ("RT-04", "Realtime", "Editor lock visible to others", "Take control in A", "B shows lock + read-only"),
    ("RT-05", "Realtime", "Editor timeline sync", "Edit timeline in A (holder)", "B's playback updates live"),
    ("RT-06", "Realtime", "Presence playheads", "Both scrub same project", "Each other's playhead visible; ghost-hide 60s"),
    ("RT-07", "Persistence", "Theme + font scale survive reload", "Set theme/text size; reload", "Prefs retained (localStorage)"),
    ("RT-08", "Persistence", "Pipeline collapse survives reload", "Collapse a stage; reload", "Stays collapsed (user_preferences)"),
    ("RT-09", "Persistence", "Hidden lanes survive reload + 2nd browser", "Hide a lane; reload; open elsewhere", "Lane hidden (synced via Supabase)"),
    ("RT-10", "Persistence", "Visited DNA rows persist", "Open some DNA rows; reload", "Visited styling retained"),
    ("RT-11", "Persistence", "Editor UI preset persists", "Set CapCut/Classic; reload", "Preset retained per-user"),
    ("RT-12", "Persistence", "Team-chat mute persists", "Mute; reload", "Stays muted"),
    ("RT-13", "Persistence", "Pipeline sub-mode persists", "Pick List/Calendar; reload", "Mode retained (wb_pipeline_mode)"),
    ("RT-14", "Persistence", "Boot survives missing pref table", "Simulate a missing user_preferences read", "App still boots (separate effect)"),
]

INTEG_HEADERS = [
    "Test ID", "Integration", "Entry point", "Connect / expected",
    "Error / reconnect path", "End-to-end smoke", "Status", "Last checked", "Notes",
]
INTEG_ROWS = [
    ("IN-01", "Facebook OAuth", "Analytics / social panel", "Connect → api.footagebrain.com/api/auth/facebook/login; token on Hetzner", "'Reconnect' link on expiry/error", "Connected badge + follower count shows"),
    ("IN-02", "Instagram insights", "Analytics (rides FB page token)", "Owner clicks Connect once; insights via FB page", "Mock fallback when unconnected", "Impressions/engagement aggregate"),
    ("IN-03", "YouTube data", "Analytics (when connected)", "OAuth api.footagebrain.com/api/auth/youtube", "Reconnect on token error", "Views/retention + inbox comments load"),
    ("IN-04", "Rocket.Chat", "Team tab (chat.footagebrain.com)", "iframe loads + WebSocket stays alive", "Poll fallback every 20s for messages", "Send share → card in channel + comment saved"),
    ("IN-05", "Planable push", "Export → Push to Planable", "DRAFT posts (never auto-publish); media via public URL", "Text-only fallback if media fails", "Drafts created; campaign id returned"),
    ("IN-06", "Epidemic music", "Music Library", "Server-side proxy (token never in browser)", "Calibration needed if private API 000s", "Search → preview → licensed download URL"),
    ("IN-07", "Scout backend", "Monitor → Scout → Refresh", "Fire-and-forget scrape via SCOUT_BACKEND_URL", "Reload re-fetches; classify on next scrape", "New products appear after ~2min"),
]

BUG_HEADERS = [
    "Bug ID", "Title", "Area / Sheet", "Severity", "Priority", "Steps to repro",
    "Expected", "Actual", "Screenshot / link", "Status", "Found by",
    "Date found", "Fixed in (commit)", "Date fixed", "Notes",
]
BUG_WIDTHS = [9, 30, 18, 11, 9, 40, 32, 32, 18, 14, 12, 12, 16, 12, 30]
BUG_STATUS_OPTS = '"Open,In-progress,Fixed,Won-t-fix,Verified"'


# --------------------------------------------------------------------------- #
# Special sheet builders
# --------------------------------------------------------------------------- #
def build_readme(wb):
    ws = wb.create_sheet("README — How to use")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    def line(row, text, font=None, fill=None):
        c = ws.cell(row=row, column=2, value=text)
        c.alignment = WRAP_TOP
        if font: c.font = font
        if fill: c.fill = fill
    line(1, "FootageBrain — Full QA / Debugging Tracker", TITLE_FONT)
    line(2, "One sheet per tab/feature. Pre-populated with concrete test cases pulled from the codebase.", SUB_FONT)
    rows = [
        (4,  "HOW TO USE", SECTION_FONT, SECTION_FILL),
        (5,  "1. Pick a feature sheet (tabs along the bottom). Work it top to bottom."),
        (6,  "2. For each row: do the Steps, compare to the Expected result."),
        (7,  "3. Set Status (dropdown col H). It auto color-codes: green Pass, red Fail, amber Blocked/Retest, grey N-A."),
        (8,  "4. On a Fail: set Severity (col I), describe what happened in Actual result / Notes (col J),"),
        (9,  "     then add a row in the Bug Log and put its BUG-id in col K (Bug ID)."),
        (10, "5. Fill Tester (L) and Date tested (M)."),
        (11, "6. The Dashboard sheet rolls up counts automatically across every feature sheet."),
        (13, "RECOMMENDED TEST ORDER", SECTION_FONT, SECTION_FILL),
        (14, "A. Smoke (owner): open every tab once — confirm each renders, no console errors / blank screens."),
        (15, "B. Per-area deep pass: walk each feature sheet."),
        (16, "C. Per-role pass: use the 'Per-Role Access Matrix' sheet. Do a quick pass via owner perspective-switch,"),
        (17, "     then confirm with REAL logins (perspective-switch can mask data-level gating)."),
        (18, "D. Realtime & persistence pass: 'Realtime & Persistence' sheet (two browsers / reload)."),
        (19, "E. Integration pass: 'Integration Health' sheet (each external connection end-to-end)."),
        (20, "F. Re-test: filter Bug Log to Fixed, re-run those rows, flip feature Status to Pass or back to Fail."),
        (22, "WHAT TO PUT IN NOTES (col J)", SECTION_FONT, SECTION_FILL),
        (23, "Exact symptom · what you clicked · console error text (if any) · the URL/?query · which role + browser · is it reproducible."),
        (24, "Enough that a fix can start from the row alone."),
        (26, "LEGEND", SECTION_FONT, SECTION_FILL),
        (27, "Priority:  P0 = blocks core use / data loss · P1 = major feature broken · P2 = important · P3 = minor / cosmetic."),
        (28, "Status:    Not started · Pass · Fail · Blocked (can't test yet) · N-A (not applicable) · Retest (after a fix)."),
        (29, "Severity:  Critical (data loss / crash) · Major (feature unusable) · Minor (workaround exists) · Cosmetic."),
        (30, "Roles:     owner · skilled · variant · reviewer · demo · all.  Slots: paul=owner, alex=Judy(skilled), sam=Jay(variant), maya=Leroy(reviewer)."),
    ]
    for tup in rows:
        if len(tup) == 4:
            line(tup[0], tup[1], tup[2], tup[3])
        else:
            line(tup[0], tup[1])
    ws.sheet_view.showGridLines = False
    return ws


def build_dashboard(wb, feature_sheets):
    """feature_sheets = list of (sheet_title, last_row)."""
    ws = wb.create_sheet("Dashboard")
    ws.cell(row=1, column=1, value="QA Progress Dashboard").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Auto-rolls up across every feature sheet. Open in Excel/Sheets to see live totals.").font = SUB_FONT

    statuses = ["Not started", "Pass", "Fail", "Blocked", "N-A", "Retest"]
    # Overall summary block
    ws.cell(row=4, column=1, value="OVERALL").font = Font(bold=True, size=12)
    headers = ["Metric", "Count"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT

    # Build COUNTIF fragments across all feature sheets' status column (H2:H<last>)
    def countif_all(value):
        parts = []
        for (t, last) in feature_sheets:
            parts.append(f"COUNTIF('{t}'!$H$2:$H${last},\"{value}\")")
        return "=" + "+".join(parts)

    def counta_all():
        parts = []
        for (t, last) in feature_sheets:
            parts.append(f"COUNTA('{t}'!$A$2:$A${last})")
        return "=" + "+".join(parts)

    ws.cell(row=6, column=1, value="Total test cases")
    ws.cell(row=6, column=2, value=counta_all())
    row = 7
    for st in statuses:
        ws.cell(row=row, column=1, value=st)
        ws.cell(row=row, column=2, value=countif_all(st))
        row += 1
    ws.cell(row=row, column=1, value="% complete (not 'Not started')").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=IF(B6=0,0,1-B7/B6)")
    ws.cell(row=row, column=2).number_format = "0%"

    # Per-area table
    start = row + 3
    ws.cell(row=start - 1, column=1, value="BY AREA").font = Font(bold=True, size=12)
    area_headers = ["Sheet", "Total", "Pass", "Fail", "Blocked", "Not started"]
    for i, h in enumerate(area_headers, 1):
        c = ws.cell(row=start, column=i, value=h); c.fill = HEADER_FILL; c.font = HEADER_FONT
    r = start + 1
    for (t, last) in feature_sheets:
        ws.cell(row=r, column=1, value=t)
        ws.cell(row=r, column=2, value=f"=COUNTA('{t}'!$A$2:$A${last})")
        ws.cell(row=r, column=3, value=f"=COUNTIF('{t}'!$H$2:$H${last},\"Pass\")")
        ws.cell(row=r, column=4, value=f"=COUNTIF('{t}'!$H$2:$H${last},\"Fail\")")
        ws.cell(row=r, column=5, value=f"=COUNTIF('{t}'!$H$2:$H${last},\"Blocked\")")
        ws.cell(row=r, column=6, value=f"=COUNTIF('{t}'!$H$2:$H${last},\"Not started\")")
        r += 1

    # Bug log rollup
    bug_start = r + 2
    ws.cell(row=bug_start - 1, column=1, value="BUG LOG").font = Font(bold=True, size=12)
    bug_metrics = [
        ("Total bugs", "=COUNTA('Bug Log'!$A$2:$A$1000)"),
        ("Open", "=COUNTIF('Bug Log'!$J$2:$J$1000,\"Open\")"),
        ("In-progress", "=COUNTIF('Bug Log'!$J$2:$J$1000,\"In-progress\")"),
        ("Fixed", "=COUNTIF('Bug Log'!$J$2:$J$1000,\"Fixed\")"),
        ("Verified", "=COUNTIF('Bug Log'!$J$2:$J$1000,\"Verified\")"),
        ("Critical (any status)", "=COUNTIF('Bug Log'!$D$2:$D$1000,\"Critical\")"),
    ]
    cb = bug_start
    for label, formula in bug_metrics:
        ws.cell(row=cb, column=1, value=label)
        ws.cell(row=cb, column=2, value=formula)
        cb += 1

    ws.column_dimensions["A"].width = 34
    for col in "BCDEF":
        ws.column_dimensions[col].width = 13
    ws.sheet_view.showGridLines = False
    return ws


def build_role_matrix(wb):
    ws = wb.create_sheet("Per-Role Access Matrix")
    for i, h in enumerate(ROLE_MATRIX_HEADERS, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, ncols=len(ROLE_MATRIX_HEADERS))
    note = ws.cell(row=1, column=len(ROLE_MATRIX_HEADERS),  value="Notes")
    widths = [9, 26, 9, 9, 9, 9, 9, 9, 13, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 2
    for (cid, name, typ, o, sk, va, rv, dm) in ROLE_ROWS:
        for c, v in enumerate([cid, name, typ, o, sk, va, rv, dm, "Not started", ""], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = TOP if c != 10 else WRAP_TOP
            if 4 <= c <= 8:
                cell.alignment = Alignment(horizontal="center", vertical="top")
        r += 1
    last = r - 1
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ROLE_MATRIX_HEADERS))}{last}"
    # Legend
    leg = ws.cell(row=last + 2, column=1, value="Legend: V=Visible · H=Hidden · A=Allowed · X=Blocked (expected per permissions-catalog.js LEAN defaults). Set Status per role check; note any mismatch.")
    leg.font = SUB_FONT
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=len(ROLE_MATRIX_HEADERS))
    # Status dropdown
    dv = DataValidation(type="list", formula1=STATUS_OPTS, allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"I2:I{last}")
    rng = f"I2:I{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL, font=PASS_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL, font=FAIL_FONT))
    return ws, last


def build_simple_sheet(wb, title, headers, rows, status_col_letter, freeze="A2"):
    ws = wb.create_sheet(title[:31])
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, ncols=len(headers))
    r = 2
    for row in rows:
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
        # default Status
        r += 1
    last = r - 1
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
    return ws, last


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    build_readme(wb)

    # Feature sheets (standard schema)
    feature_defs = [
        ("Auth & Access", AUTH),
        ("Navigation & Permissions", NAV),
        ("My Work", MYWORK),
        ("Pipeline", PIPELINE),
        ("Reel DNA", REELDNA),
        ("Thumbnail DNA", THUMB),
        ("Footage Coverage Locations", LIBS),
        ("Reel Detail", DETAIL),
        ("Editor", EDITOR),
        ("Editor Collab & Projects", EDITOR_COLLAB),
        ("Export & Planable", EXPORT),
        ("Music Library", MUSIC),
        ("Scout", SCOUT),
        ("Monitor Hub", MONITOR),
        ("Team Chat & Notifications", TEAMCHAT),
        ("Training & Resources", TRAINING),
        ("Inbox Analytics Generate", MISC),
        ("Admin — Roles & Permissions", ADMIN),
    ]
    feature_sheets = []   # (title, last_row) for dashboard
    for title, rows in feature_defs:
        ws, last = add_feature_sheet(wb, title, rows)
        feature_sheets.append((ws.title, last))

    # Cross-cutting sheets
    rm_ws, rm_last = build_role_matrix(wb)
    rt_ws, rt_last = build_simple_sheet(wb, "Realtime & Persistence", REALTIME_HEADERS, REALTIME_ROWS, "F")
    # add status dropdowns to realtime
    dv = DataValidation(type="list", formula1=STATUS_OPTS, allow_blank=True)
    rt_ws.add_data_validation(dv); dv.add(f"F2:F{rt_last}")
    for w, col in zip([10, 16, 40, 40, 44, 13, 40], "ABCDEFG"):
        rt_ws.column_dimensions[col].width = w

    in_ws, in_last = build_simple_sheet(wb, "Integration Health", INTEG_HEADERS, INTEG_ROWS, "G")
    dv2 = DataValidation(type="list", formula1=STATUS_OPTS, allow_blank=True)
    in_ws.add_data_validation(dv2); dv2.add(f"G2:G{in_last}")
    for w, col in zip([9, 18, 26, 34, 30, 34, 13, 14, 30], "ABCDEFGHI"):
        in_ws.column_dimensions[col].width = w

    # Bug Log
    bug_ws = wb.create_sheet("Bug Log")
    for i, h in enumerate(BUG_HEADERS, 1):
        bug_ws.cell(row=1, column=i, value=h)
        bug_ws.column_dimensions[get_column_letter(i)].width = BUG_WIDTHS[i - 1]
    style_header(bug_ws, ncols=len(BUG_HEADERS))
    bug_ws.freeze_panes = "A2"
    bug_ws.auto_filter.ref = f"A1:{get_column_letter(len(BUG_HEADERS))}500"
    dv_sev = DataValidation(type="list", formula1=SEVERITY_OPTS, allow_blank=True)
    dv_pri = DataValidation(type="list", formula1=PRIORITY_OPTS, allow_blank=True)
    dv_bst = DataValidation(type="list", formula1=BUG_STATUS_OPTS, allow_blank=True)
    bug_ws.add_data_validation(dv_sev); bug_ws.add_data_validation(dv_pri); bug_ws.add_data_validation(dv_bst)
    dv_sev.add("D2:D500"); dv_pri.add("E2:E500"); dv_bst.add("J2:J500")
    rng = "J2:J500"
    bug_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Open"'], fill=FAIL_FILL, font=FAIL_FONT))
    bug_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"In-progress"'], fill=WARN_FILL, font=WARN_FONT))
    bug_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fixed"'], fill=PASS_FILL, font=PASS_FONT))
    bug_ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Verified"'], fill=PASS_FILL, font=PASS_FONT))

    # Dashboard last (needs feature_sheets list) — insert near front
    dash = build_dashboard(wb, feature_sheets)

    # Reorder: README, Dashboard, ...features..., role matrix, realtime, integ, bug log
    order = ["README — How to use", "Dashboard"] + [t for (t, _) in feature_sheets] + \
            ["Per-Role Access Matrix", "Realtime & Persistence", "Integration Health", "Bug Log"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 999)

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "QA-Debug-Tracker-FootageBrain.xlsx")
    wb.save(out)

    total = sum(last - 1 for (_, last) in feature_sheets)
    print(f"Wrote {out}")
    print(f"Sheets: {len(wb.sheetnames)}  ·  feature test rows: {total}")
    print("Sheet list:")
    for n in wb.sheetnames:
        print("  -", n)


if __name__ == "__main__":
    main()
